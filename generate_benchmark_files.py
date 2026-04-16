from __future__ import annotations

import argparse
import shutil
from pathlib import Path

from openpyxl import load_workbook

from app import build_dataframes, extract_scalars, solvePTO


def load_sheet_records(workbook_path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records = []
    for row in rows[1:]:
        if row is None or all(cell in (None, "") for cell in row):
            continue
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[idx] if idx < len(row) else None
        records.append(record)
    return records


def load_settings(workbook_path: Path) -> dict:
    settings = {}
    for row in load_sheet_records(workbook_path, "Settings"):
        field = row.get("field")
        if field in (None, ""):
            continue
        settings[str(field)] = row.get("value")
    return settings


def load_first_available_sheet_records(workbook_path: Path, sheet_names: list[str]) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=True)
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            records = []
            for row in rows[1:]:
                if row is None or all(cell in (None, "") for cell in row):
                    continue
                record = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    record[header] = row[idx] if idx < len(row) else None
                records.append(record)
            return records
    raise ValueError(
        f"None of the expected sheets {sheet_names} were found in {workbook_path}")


def build_input_data(workbook_path: Path, spot_prices_path: Path | None = None,
                     tariffs_path: Path | None = None) -> tuple[dict, dict]:
    settings = load_settings(workbook_path)
    trips = [
        row for row in load_sheet_records(workbook_path, "Trips")
        if row.get("time_begin") not in (None, "")
        and row.get("time_end") not in (None, "")
        and row.get("energy_kwhkm") not in (None, "")
    ]
    spot_source_path = spot_prices_path if spot_prices_path else workbook_path
    spot_prices = load_first_available_sheet_records(
        spot_source_path, ["Spot Prices", "Prices"])

    tariffs = []
    if tariffs_path:
        tariffs = load_first_available_sheet_records(tariffs_path, ["Tariffs", "Prices"])
    else:
        workbook = load_workbook(workbook_path, data_only=True)
        if "Tariffs" in workbook.sheetnames:
            tariffs = load_sheet_records(workbook_path, "Tariffs")

    input_data = {
        "timestep_minutes": settings.get("timestep_minutes"),
        "v2g_enabled": settings.get("v2g_enabled"),
        "buses": load_sheet_records(workbook_path, "Buses"),
        "chargers": load_sheet_records(workbook_path, "Chargers"),
        "trip_time": trips,
        "prices": spot_prices,
        "grid_prices": spot_prices,
        "tariffs": tariffs,
        "realtime_state": load_sheet_records(workbook_path, "Realtime state"),
    }
    return input_data, settings


def status_for_bus(model, bus_idx: int, timestep: int, trip_count: int, charger_count: int) -> tuple[str, int | None, int | None]:
    for trip_idx in range(1, trip_count + 1):
        if round(model.b[bus_idx, trip_idx, timestep].value or 0) == 1:
            return "in_trip", trip_idx, None
    for charger_idx in range(1, charger_count + 1):
        if round(model.x[bus_idx, charger_idx, timestep].value or 0) == 1:
            return "charging", None, charger_idx
    for charger_idx in range(1, charger_count + 1):
        if round(model.y[bus_idx, charger_idx, timestep].value or 0) == 1:
            return "discharging", None, charger_idx
    return "idle", None, None


def find_row_by_column_value(worksheet, column_idx: int, value) -> int | None:
    for row_idx in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row_idx, column=column_idx).value == value:
            return row_idx
    return None


def update_settings_sheet(worksheet, optimization_mode: str) -> None:
    for row_idx in range(2, worksheet.max_row + 1):
        field = worksheet.cell(row=row_idx, column=1).value
        if field == "optimization_mode":
            worksheet.cell(row=row_idx, column=2, value=optimization_mode)


def update_realtime_sheet(workbook, sc: dict, model, timestep: int) -> None:
    worksheet = workbook["Realtime state"]
    trip_count = sc["i_count"]
    charger_count = sc["n_count"]

    for bus_idx in range(1, sc["k_count"] + 1):
        row_idx = find_row_by_column_value(worksheet, 2, bus_idx)
        if row_idx is None:
            row_idx = worksheet.max_row + 1
            worksheet.cell(row=row_idx, column=2, value=bus_idx)

        status, trip_idx, charger_idx = status_for_bus(
            model, bus_idx, timestep, trip_count, charger_count)
        energy_kwh = float(model.e[bus_idx, timestep].value or 0.0)
        soc_pct = round((energy_kwh / sc["C_bat"][bus_idx - 1]) * 100.0, 3)

        worksheet.cell(row=row_idx, column=1, value=timestep)
        worksheet.cell(row=row_idx, column=2, value=bus_idx)
        worksheet.cell(row=row_idx, column=3, value=soc_pct)
        worksheet.cell(row=row_idx, column=4, value=round(energy_kwh, 6))
        worksheet.cell(row=row_idx, column=5, value=status)
        worksheet.cell(row=row_idx, column=6, value=0)


def write_benchmark_action_sheet(workbook, sc: dict, model, timestep: int) -> None:
    if "Benchmark action" in workbook.sheetnames:
        del workbook["Benchmark action"]

    worksheet = workbook.create_sheet("Benchmark action")
    worksheet.append([
        "current_timestep",
        "bus_id",
        "operation_status",
        "trip_id",
        "charger_id",
        "energy_kwh",
        "soc_percent",
        "site_buy_kwh",
        "site_sell_kwh",
    ])

    site_buy = float(model.w_buy[timestep].value or 0.0)
    site_sell = float(model.w_sell[timestep].value or 0.0)
    trip_count = sc["i_count"]
    charger_count = sc["n_count"]

    for bus_idx in range(1, sc["k_count"] + 1):
        status, trip_idx, charger_idx = status_for_bus(
            model, bus_idx, timestep, trip_count, charger_count)
        energy_kwh = float(model.e[bus_idx, timestep].value or 0.0)
        soc_pct = round((energy_kwh / sc["C_bat"][bus_idx - 1]) * 100.0, 3)
        worksheet.append([
            timestep,
            bus_idx,
            status,
            trip_idx,
            charger_idx,
            round(energy_kwh, 6),
            soc_pct,
            round(site_buy, 6),
            round(site_sell, 6),
        ])


def generate_benchmark_files(template_path: Path, output_dir: Path,
                             spot_prices_path: Path | None = None,
                             tariffs_path: Path | None = None) -> Path:
    input_data, settings = build_input_data(
        template_path,
        spot_prices_path=spot_prices_path,
        tariffs_path=tariffs_path,
    )
    price_guidance = {}
    data = build_dataframes(input_data)
    sc = extract_scalars(
        data,
        price_guidance=price_guidance,
        optimization_mode="day_ahead",
        current_timestep=1,
    )
    model = solvePTO(sc)
    if model is None:
        raise RuntimeError("Day-ahead optimization was infeasible; benchmark files were not generated.")

    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    summary_path = output_dir / "benchmark_summary.txt"
    summary_lines = [
        f"source_template={template_path}",
        f"timestep_minutes={sc['timestep_minutes']}",
        f"optimized_steps={sc['T_steps']}",
        f"bus_count={sc['k_count']}",
        f"charger_count={sc['n_count']}",
    ]

    for timestep in range(1, sc["T_steps"] + 1):
        workbook = load_workbook(template_path)
        update_settings_sheet(workbook["Settings"], "real_time")
        update_realtime_sheet(workbook, sc, model, timestep)
        write_benchmark_action_sheet(workbook, sc, model, timestep)

        output_path = output_dir / f"benchmark_timestep_{timestep:02d}.xlsx"
        workbook.save(output_path)
        summary_lines.append(str(output_path.name))

    summary_path.write_text("\n".join(summary_lines) + "\n")
    return output_dir


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate rolling benchmark Excel files from the day-ahead optimization.")
    parser.add_argument(
        "--input",
        default="Files/Inputs.xlsx",
        help="Path to the cleaned input workbook template.",
    )
    parser.add_argument(
        "--output-dir",
        default="Files/day_ahead_benchmark",
        help="Folder where the 48 benchmark workbooks will be written.",
    )
    parser.add_argument(
        "--spot-prices-file",
        default="",
        help="Optional Excel file containing spot prices (sheet 'Spot Prices' or 'Prices').",
    )
    parser.add_argument(
        "--tariffs-file",
        default="",
        help="Optional Excel file containing buy/sell tariffs (sheet 'Tariffs' or 'Prices').",
    )
    args = parser.parse_args()

    spot_prices_path = Path(args.spot_prices_file) if args.spot_prices_file else None
    tariffs_path = Path(args.tariffs_file) if args.tariffs_file else None

    output_dir = generate_benchmark_files(
        template_path=Path(args.input),
        output_dir=Path(args.output_dir),
        spot_prices_path=spot_prices_path,
        tariffs_path=tariffs_path,
    )
    print(f"Generated benchmark files in {output_dir}")


if __name__ == "__main__":
    main()
