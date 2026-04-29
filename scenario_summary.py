from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean

import pyomo.environ as pyo
from openpyxl import Workbook, load_workbook

SUMMARY_HEADERS = [
    'run_timestamp_utc',
    'scenario',
    'input_workbook',
    'spot_prices_file',
    'tariffs_file',
    'optimization_mode',
    'v2g_enabled',
    'timestep_minutes',
    'optimized_steps',
    'bus_count',
    'charger_count',
    'trip_count',
    'avg_grid_price',
    'avg_buy_price',
    'avg_sell_price',
    'buy_multipliers',
    'sell_multipliers',
    'period_boundaries',
    'pto_daily_cost',
    'aggregator_revenue',
    'aggregator_buy_margin',
    'aggregator_sell_margin',
    'total_buy_cost',
    'total_sell_revenue',
    'net_daily_cost',
    'total_kwh_bought',
    'total_kwh_sold',
    'peak_grid_import_kw',
    'peak_grid_export_kw',
    'total_trip_energy_kwh',
    'cost_per_kwh_served',
    'minimum_soc_before_departure',
    'average_soc_before_departure',
    'number_of_risked_departures',
    'trip_feasibility_rate',
    'end_of_day_average_soc',
    'end_of_day_soc_by_bus',
    'share_of_charging_in_low_price_hours',
    'share_of_charging_in_high_price_hours',
    'battery_throughput_kwh',
    'delta_vs_dumb_cost',
    'delta_vs_dumb_cost_pct',
]

REASONING_HEADERS = [
    'run_timestamp_utc',
    'scenario',
    'reasoning_source',
    'reasoning_text',
    'input_workbook',
    'spot_prices_file',
    'tariffs_file',
    'optimization_mode',
    'v2g_enabled',
    'avg_grid_price',
    'avg_buy_price',
    'avg_sell_price',
    'buy_multipliers',
    'sell_multipliers',
    'period_boundaries',
    'total_kwh_bought',
    'total_kwh_sold',
    'peak_grid_import_kw',
    'peak_grid_export_kw',
    'share_of_charging_in_low_price_hours',
    'share_of_charging_in_high_price_hours',
]


def _value(expr) -> float:
    value = pyo.value(expr)
    return float(value) if value is not None else 0.0


def _json_text(value) -> str:
    return json.dumps(value)


def _quantile_threshold(values: list[float], fraction: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(float(v) for v in values)
    index = int(round((len(ordered) - 1) * fraction))
    index = max(0, min(len(ordered) - 1, index))
    return ordered[index]


def _end_of_day_energy(model, sc: dict, bus_idx: int) -> float:
    t_final = sc['T_steps']
    return _value(model.e[bus_idx, t_final])


def _departure_soc_metrics(sc: dict, model) -> tuple[list[float], int]:
    departure_socs = []
    risked_departures = 0
    for trip_idx in range(1, sc['i_count'] + 1):
        bus_idx = trip_idx
        if bus_idx > sc['k_count']:
            break
        departure_t = max(1, min(sc['T_steps'], sc['T_start'][trip_idx - 1]))
        departure_energy = _value(model.e[bus_idx, departure_t])
        capacity = float(sc['C_bat'][bus_idx - 1])
        departure_soc = departure_energy / capacity if capacity else 0.0
        departure_socs.append(departure_soc)

        trip_duration = max(0, sc['T_end'][trip_idx - 1] - sc['T_start'][trip_idx - 1])
        required_energy = (sc['gama'][trip_idx - 1] * trip_duration) + (sc['E_min'] * capacity)
        if departure_energy + 1e-9 < required_energy:
            risked_departures += 1
    return departure_socs, risked_departures


def build_day_ahead_summary_row(
    sc: dict,
    model,
    scenario: str,
    *,
    input_workbook: Path | None = None,
    spot_prices_file: Path | None = None,
    tariffs_file: Path | None = None,
) -> dict:
    t_steps = sc['T_steps']
    timestep_hours = sc['timestep_minutes'] / 60.0

    w_buy = [_value(model.w_buy[t]) for t in range(1, t_steps + 1)]
    w_sell = [_value(model.w_sell[t]) for t in range(1, t_steps + 1)]
    total_buy_cost = sum(sc['S_buy'][t - 1] * w_buy[t - 1] for t in range(1, t_steps + 1))
    total_sell_revenue = sum(sc['S_sell'][t - 1] * w_sell[t - 1] for t in range(1, t_steps + 1))
    net_daily_cost = total_buy_cost - total_sell_revenue
    aggregator_buy_margin = sum((sc['S_buy'][t - 1] - sc['P'][t - 1]) * w_buy[t - 1] for t in range(1, t_steps + 1))
    aggregator_sell_margin = sum((sc['P'][t - 1] - sc['S_sell'][t - 1]) * w_sell[t - 1] for t in range(1, t_steps + 1))
    aggregator_revenue = aggregator_buy_margin + aggregator_sell_margin
    total_kwh_bought = sum(w_buy)
    total_kwh_sold = sum(w_sell)

    total_trip_energy = sum(
        sc['gama'][idx] * max(0, sc['T_end'][idx] - sc['T_start'][idx])
        for idx in range(sc['i_count'])
    )
    peak_grid_import_kw = (max(w_buy) / timestep_hours) if (w_buy and timestep_hours) else 0.0
    peak_grid_export_kw = (max(w_sell) / timestep_hours) if (w_sell and timestep_hours) else 0.0
    cost_per_kwh_served = (net_daily_cost / total_trip_energy) if total_trip_energy else None

    departure_socs, risked_departures = _departure_soc_metrics(sc, model)
    minimum_soc_before_departure = (min(departure_socs) * 100.0) if departure_socs else None
    average_soc_before_departure = (mean(departure_socs) * 100.0) if departure_socs else None
    trip_feasibility_rate = (
        (len(departure_socs) - risked_departures) / len(departure_socs)
        if departure_socs else None
    )

    end_of_day_soc_by_bus = []
    for bus_idx in range(1, sc['k_count'] + 1):
        capacity = float(sc['C_bat'][bus_idx - 1])
        ending_energy = _end_of_day_energy(model, sc, bus_idx)
        end_of_day_soc_by_bus.append((ending_energy / capacity) * 100.0 if capacity else 0.0)
    end_of_day_average_soc = mean(end_of_day_soc_by_bus) if end_of_day_soc_by_bus else None

    low_threshold = _quantile_threshold(sc['P'], 0.25)
    high_threshold = _quantile_threshold(sc['P'], 0.75)
    low_price_charge = sum(w for w, p in zip(w_buy, sc['P']) if p <= low_threshold)
    high_price_charge = sum(w for w, p in zip(w_buy, sc['P']) if p >= high_threshold)
    share_low_price = (low_price_charge / total_kwh_bought) if total_kwh_bought else None
    share_high_price = (high_price_charge / total_kwh_bought) if total_kwh_bought else None

    return {
        'run_timestamp_utc': datetime.now(timezone.utc).isoformat(),
        'scenario': scenario,
        'input_workbook': str(input_workbook) if input_workbook else '',
        'spot_prices_file': str(spot_prices_file) if spot_prices_file else '',
        'tariffs_file': str(tariffs_file) if tariffs_file else '',
        'optimization_mode': sc['optimization_mode'],
        'v2g_enabled': bool(sc.get('v2g_enabled', True)),
        'timestep_minutes': sc['timestep_minutes'],
        'optimized_steps': t_steps,
        'bus_count': sc['k_count'],
        'charger_count': sc['n_count'],
        'trip_count': sc['i_count'],
        'avg_grid_price': sc['avg_P'],
        'avg_buy_price': sc['avg_S_buy'],
        'avg_sell_price': sc['avg_S_sell'],
        'buy_multipliers': _json_text(sc['buy_multipliers']),
        'sell_multipliers': _json_text(sc['sell_multipliers']),
        'period_boundaries': _json_text(sc['boundaries']),
        'pto_daily_cost': net_daily_cost,
        'aggregator_revenue': aggregator_revenue,
        'aggregator_buy_margin': aggregator_buy_margin,
        'aggregator_sell_margin': aggregator_sell_margin,
        'total_buy_cost': total_buy_cost,
        'total_sell_revenue': total_sell_revenue,
        'net_daily_cost': net_daily_cost,
        'total_kwh_bought': total_kwh_bought,
        'total_kwh_sold': total_kwh_sold,
        'peak_grid_import_kw': peak_grid_import_kw,
        'peak_grid_export_kw': peak_grid_export_kw,
        'total_trip_energy_kwh': total_trip_energy,
        'cost_per_kwh_served': cost_per_kwh_served,
        'minimum_soc_before_departure': minimum_soc_before_departure,
        'average_soc_before_departure': average_soc_before_departure,
        'number_of_risked_departures': risked_departures,
        'trip_feasibility_rate': trip_feasibility_rate,
        'end_of_day_average_soc': end_of_day_average_soc,
        'end_of_day_soc_by_bus': _json_text([round(v, 4) for v in end_of_day_soc_by_bus]),
        'share_of_charging_in_low_price_hours': share_low_price,
        'share_of_charging_in_high_price_hours': share_high_price,
        'battery_throughput_kwh': total_kwh_bought + total_kwh_sold,
        'delta_vs_dumb_cost': None,
        'delta_vs_dumb_cost_pct': None,
    }



def _base_flow_metrics(sc: dict, model) -> dict:
    t_steps = sc['T_steps']
    timestep_hours = sc['timestep_minutes'] / 60.0
    w_buy = [_value(model.w_buy[t]) for t in range(1, t_steps + 1)]
    w_sell = [_value(model.w_sell[t]) for t in range(1, t_steps + 1)]
    total_kwh_bought = sum(w_buy)
    total_kwh_sold = sum(w_sell)
    low_threshold = _quantile_threshold(sc['P'], 0.25)
    high_threshold = _quantile_threshold(sc['P'], 0.75)
    low_price_charge = sum(w for w, p in zip(w_buy, sc['P']) if p <= low_threshold)
    high_price_charge = sum(w for w, p in zip(w_buy, sc['P']) if p >= high_threshold)
    return {
        'w_buy': w_buy,
        'w_sell': w_sell,
        'total_kwh_bought': total_kwh_bought,
        'total_kwh_sold': total_kwh_sold,
        'peak_grid_import_kw': (max(w_buy) / timestep_hours) if (w_buy and timestep_hours) else 0.0,
        'peak_grid_export_kw': (max(w_sell) / timestep_hours) if (w_sell and timestep_hours) else 0.0,
        'share_low_price': (low_price_charge / total_kwh_bought) if total_kwh_bought else None,
        'share_high_price': (high_price_charge / total_kwh_bought) if total_kwh_bought else None,
    }


def _auto_reasoning_text(scenario: str, sc: dict, metrics: dict) -> tuple[str, str]:
    share_low = metrics['share_low_price']
    share_high = metrics['share_high_price']
    low_pct = f"{(share_low * 100.0):.1f}%" if share_low is not None else 'n/a'
    high_pct = f"{(share_high * 100.0):.1f}%" if share_high is not None else 'n/a'

    if scenario == 'dumb_charging_no_v2g':
        return (
            'deterministic_rule',
            'Rule-based benchmark without V2G. The objective charges buses as early and as much as possible before departures, using spot-market tariffs directly instead of responding to a learned tariff policy. '
            f'The resulting schedule bought {metrics["total_kwh_bought"]:.3f} kWh, sent {low_pct} of charging to low-price hours, and ignored export opportunities.'
        )

    if scenario == 'optimization_no_v2g':
        return (
            'deterministic_optimization',
            'Cost-minimizing smart-charging benchmark without V2G. The solver uses spot-market tariffs directly, shifts charging toward cheaper periods when feasible, and disables discharge decisions. '
            f'The resulting schedule bought {metrics["total_kwh_bought"]:.3f} kWh, allocated {low_pct} of charging to low-price hours, and {high_pct} to high-price hours.'
        )

    return (
        'aggregator_policy',
        'Aggregator-managed smart charging with V2G enabled. The solver minimizes PTO net charging cost under the applied tariff schedule while allowing exports when the tariff spread and SOC constraints make V2G attractive. '
        f'The resulting schedule bought {metrics["total_kwh_bought"]:.3f} kWh, sold {metrics["total_kwh_sold"]:.3f} kWh, placed {low_pct} of charging in low-price hours, and reached a peak export of {metrics["peak_grid_export_kw"]:.3f} kW.'
    )


def build_agent_reasoning_row(
    sc: dict,
    model,
    scenario: str,
    *,
    input_workbook: Path | None = None,
    spot_prices_file: Path | None = None,
    tariffs_file: Path | None = None,
    reasoning_source: str | None = None,
    reasoning_text: str | None = None,
) -> dict:
    metrics = _base_flow_metrics(sc, model)
    auto_source, auto_text = _auto_reasoning_text(scenario, sc, metrics)
    return {
        'run_timestamp_utc': datetime.now(timezone.utc).isoformat(),
        'scenario': scenario,
        'reasoning_source': reasoning_source or auto_source,
        'reasoning_text': reasoning_text or auto_text,
        'input_workbook': str(input_workbook) if input_workbook else '',
        'spot_prices_file': str(spot_prices_file) if spot_prices_file else '',
        'tariffs_file': str(tariffs_file) if tariffs_file else '',
        'optimization_mode': sc['optimization_mode'],
        'v2g_enabled': bool(sc.get('v2g_enabled', True)),
        'avg_grid_price': sc['avg_P'],
        'avg_buy_price': sc['avg_S_buy'],
        'avg_sell_price': sc['avg_S_sell'],
        'buy_multipliers': _json_text(sc['buy_multipliers']),
        'sell_multipliers': _json_text(sc['sell_multipliers']),
        'period_boundaries': _json_text(sc['boundaries']),
        'total_kwh_bought': metrics['total_kwh_bought'],
        'total_kwh_sold': metrics['total_kwh_sold'],
        'peak_grid_import_kw': metrics['peak_grid_import_kw'],
        'peak_grid_export_kw': metrics['peak_grid_export_kw'],
        'share_of_charging_in_low_price_hours': metrics['share_low_price'],
        'share_of_charging_in_high_price_hours': metrics['share_high_price'],
    }


def append_agent_reasoning(workbook_path: Path, row: dict) -> Path:
    workbook_path = Path(workbook_path)
    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()

    worksheet = _get_or_create_sheet(workbook, 'agent_reasoning', REASONING_HEADERS)

    if 'Sheet' in workbook.sheetnames and workbook['Sheet'].max_row == 1 and workbook['Sheet'].max_column == 1 and workbook['Sheet']['A1'].value is None:
        del workbook['Sheet']

    headers = _ensure_sheet_headers(worksheet, REASONING_HEADERS)
    worksheet.append([row.get(header) for header in headers])
    workbook.save(workbook_path)
    return workbook_path

def _get_or_create_sheet(workbook, sheet_name: str, headers: list[str]):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(headers)
    return worksheet


def _ensure_sheet_headers(worksheet, required_headers: list[str]) -> list[str]:
    if worksheet.max_row == 0:
        worksheet.append(required_headers)
        return required_headers[:]

    first_row = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
    existing_headers = [str(value).strip() if value not in (None, '') else '' for value in first_row]
    if not any(existing_headers):
        existing_headers = required_headers[:]
        for col_idx, header in enumerate(existing_headers, start=1):
            worksheet.cell(row=1, column=col_idx, value=header)
        return existing_headers

    for header in required_headers:
        if header not in existing_headers:
            existing_headers.append(header)
            worksheet.cell(row=1, column=len(existing_headers), value=header)
    return existing_headers


def _ensure_headers(worksheet) -> list[str]:
    return _ensure_sheet_headers(worksheet, SUMMARY_HEADERS)


def _find_latest_dumb_cost(worksheet, headers: list[str]) -> float | None:
    if 'scenario' not in headers or 'net_daily_cost' not in headers:
        return None
    scenario_col = headers.index('scenario') + 1
    cost_col = headers.index('net_daily_cost') + 1
    for row_idx in range(worksheet.max_row, 1, -1):
        if worksheet.cell(row=row_idx, column=scenario_col).value != 'dumb_charging_no_v2g':
            continue
        value = worksheet.cell(row=row_idx, column=cost_col).value
        if value in (None, ''):
            continue
        return float(value)
    return None


def append_day_ahead_summary(workbook_path: Path, row: dict) -> Path:
    workbook_path = Path(workbook_path)
    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()
    worksheet = _get_or_create_sheet(workbook, 'day_ahead_summary', SUMMARY_HEADERS)

    if 'Sheet' in workbook.sheetnames and workbook['Sheet'].max_row == 1 and workbook['Sheet'].max_column == 1 and workbook['Sheet']['A1'].value is None:
        del workbook['Sheet']

    headers = _ensure_headers(worksheet)
    baseline_cost = _find_latest_dumb_cost(worksheet, headers)
    scenario = row.get('scenario')
    if scenario == 'dumb_charging_no_v2g':
        row['delta_vs_dumb_cost'] = 0.0
        row['delta_vs_dumb_cost_pct'] = 0.0
    elif baseline_cost is not None and row.get('net_daily_cost') is not None:
        delta = float(row['net_daily_cost']) - baseline_cost
        row['delta_vs_dumb_cost'] = delta
        row['delta_vs_dumb_cost_pct'] = (delta / baseline_cost) if baseline_cost else None

    worksheet.append([row.get(header) for header in headers])
    workbook.save(workbook_path)
    return workbook_path
